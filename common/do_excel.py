# -*- coding:utf-8 _*-
import openpyxl

class Case:
    """
    测试用例封装类
    """

    def __init__(self):
        self.case_id = None
        self.url = None
        self.data = None
        self.title = None
        self.method = None
        self.expected = None
        self.actual = None
        self.result = None


class DoExcel:
    """
    excel 操作封装类
    """

    def __init__(self, file_name):  # DoExcel初始化函数
        try:
            self.file_name = file_name
            # 打开一个excel文件，返回一个workbook对象实例，把它定义为DoExcel的属性，以便于在这个类的其他地方使用
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            # 文件未找到异常处理
            print('{0} not found, please check file path'.format(file_name))
            raise e

    def get_cases(self, sheet_name):  # 根据sheet名称，获取在这个sheet里面的所有测试用例数据
        sheet = self.workbook[sheet_name]  # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row  # 获取sheet最大行数
        cases = []  # 定义一个列表，用来存放即将要放进去的测试用例
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case = Case()  # 实例化一个case对象，用来存放测试数据
            case.case_id = sheet.cell(row=r, column=1).value  # 取第r行，第1格的值
            case.title = sheet.cell(row=r, column=2).value  # 取第r行，第2格的值
            case.method = sheet.cell(row=r, column=3).value# 取第r行，第3格的值
            case.url = sheet.cell(row=r, column=4).value  # 取第r行，第4格的值
            case.data = sheet.cell(row=r, column=5).value  # 取第r行，第5格的值
            case.expected = sheet.cell(row=r, column=6).value  # 取第r行，第6格的值
            case.extract = sheet.cell(row=r, column=9).value
            cases.append(case)  # 将case放到cases 列表里面

        return cases  # for 循环结束后返回cases列表

    def get_sheet_names(self):  # 获取到work boot里面所有的sheet名称的列表
        return self.workbook.sheetnames

    # 根据sheet name 定位到sheet，然后根据case_id定位到行，取到当前行里面actual这个单元格，
    # 然后给他赋值，最后保存当前workbook
    def write_back_by_case_id(self, sheet_name, case_id, actual, result):
        sheet = self.workbook[sheet_name]  # 根据sheet名称获取sheet对象实例
        max_row = sheet.max_row  # 获取sheet最大行数
        for r in range(2, max_row + 1):  # for 循环，从第二行开始遍历
            case_id_r = sheet.cell(r, 1).value  # 获取第r行，第1列，也就是case_id这一列的值
            if case_id_r == case_id:  # 判断excel里面取到的当前行的case id 是否等于传进来的case id
                sheet.cell(r, 7).value = actual  # 写入传进来的actual到当前行的actual列的单元格
                sheet.cell(r, 8).value = result  # 写入传进来的result到当前行的result列的单元格
                self.workbook.save(filename=self.file_name)
                break


